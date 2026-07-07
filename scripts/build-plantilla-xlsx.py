# Genera public/recursos/downloads/plantilla-cuadrante-turnos-shiftia.xlsx
# Plantilla mensual de cuadrante de turnos con la marca Shiftia:
# hoja de instrucciones + hoja de cuadrante con fórmulas de recuento por
# persona (noches, findes, libres), cobertura por día/turno y chequeo de
# mínimos. Ejecutar con: python3 scripts/build-plantilla-xlsx.py
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter

INK = '0E0F0F'; PAPER = 'FAF9F6'; TEAL = '0F7A6D'; TEAL_D = '0A5950'
BORDER_C = 'E5E2DA'; RED = 'A31C22'; AMBER = '8A6220'
NIGHT_FILL = PatternFill('solid', fgColor='E4E1F5')
REST_FILL = PatternFill('solid', fgColor='EFEDE6')
HDR_FILL = PatternFill('solid', fgColor=INK)
SOFT_FILL = PatternFill('solid', fgColor='E9F5F2')
thin = Side(style='thin', color=BORDER_C)
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

DAYS = 31
NAMES = 14  # filas de personas
FIRST_DAY_COL = 3  # C
LAST_DAY_COL = FIRST_DAY_COL + DAYS - 1  # AG
FIRST_PERSON_ROW = 7
LAST_PERSON_ROW = FIRST_PERSON_ROW + NAMES - 1

wb = Workbook()

# ===== Hoja 1: Instrucciones =====
info = wb.active
info.title = 'Léeme'
info.sheet_view.showGridLines = False
info.column_dimensions['A'].width = 4
info.column_dimensions['B'].width = 96
rows = [
    ('title', 'Plantilla de cuadrante de turnos · Shiftia'),
    ('text', 'Cuadrante mensual con recuentos automáticos de noches, fines de semana y libranzas por persona, y chequeo de cobertura por día.'),
    ('gap', ''),
    ('h', 'Cómo usarla'),
    ('text', '1 · En la hoja "Cuadrante", pon el año en B2 y el mes (1-12) en B3: las fechas y los findes se pintan solos.'),
    ('text', '2 · Escribe los nombres de tu equipo en la columna A (hay 14 filas; inserta más si lo necesitas copiando una fila entera).'),
    ('text', '3 · Rellena cada día con un código: M (mañana), T (tarde), N (noche), L (libre), V (vacaciones), B (baja).'),
    ('text', '4 · En la fila "Mínimo por turno" indica cuánta gente necesitas de M, T y N: la fila de cobertura se pone roja si un día no llegas.'),
    ('gap', ''),
    ('h', 'Lo que la plantilla NO puede vigilar'),
    ('text', 'Las 12 h de descanso entre turnos (art. 34.3 ET), el descanso semanal de 36 h (art. 37.1 ET), la equidad real del reparto o las reglas de tu convenio: eso Excel no lo ve, y es donde suelen caer las sanciones y las quejas.'),
    ('text', 'Para eso tienes dos opciones gratuitas: la auditoría de tu cuadrante (súbelo y recibes un informe PDF con los incumplimientos señalados) en www.shiftia.es/recursos/auditoria-cuadrante — o dejar que Shiftia genere el cuadrante ya correcto: www.shiftia.es'),
    ('gap', ''),
    ('small', 'Plantilla gratuita de Shiftia (www.shiftia.es) · %s · Puedes compartirla libremente.' % datetime.date.today().strftime('%B %Y')),
]
r = 2
for kind, text in rows:
    c = info.cell(row=r, column=2, value=text)
    if kind == 'title':
        c.font = Font(name='Georgia', size=20, bold=False, color=INK)
    elif kind == 'h':
        c.font = Font(name='Calibri', size=13, bold=True, color=TEAL_D)
    elif kind == 'small':
        c.font = Font(name='Calibri', size=9, color='8A8A85')
    else:
        c.font = Font(name='Calibri', size=11, color='333333')
        c.alignment = Alignment(wrap_text=True, vertical='top')
        if len(text) > 90:
            info.row_dimensions[r].height = 30
    r += 1

# ===== Hoja 2: Cuadrante =====
ws = wb.create_sheet('Cuadrante')
ws.sheet_view.showGridLines = False
ws.freeze_panes = 'C7'
ws.column_dimensions['A'].width = 22
ws.column_dimensions['B'].width = 9
for col in range(FIRST_DAY_COL, LAST_DAY_COL + 1):
    ws.column_dimensions[get_column_letter(col)].width = 4.2

# Cabecera de marca
ws.merge_cells('A1:F1')
c = ws['A1']; c.value = 'Shiftia · Cuadrante mensual'
c.font = Font(name='Georgia', size=16, italic=True, color=INK)
ws['A2'] = 'Año'; ws['B2'] = datetime.date.today().year
ws['A3'] = 'Mes (1-12)'; ws['B3'] = datetime.date.today().month
for rc in ('A2', 'A3'):
    ws[rc].font = Font(bold=True, size=10, color=TEAL_D)
for rc in ('B2', 'B3'):
    ws[rc].font = Font(size=10); ws[rc].fill = SOFT_FILL; ws[rc].border = BOX
ws.merge_cells('D2:N3')
leg = ws['D2']
leg.value = 'Códigos: M mañana · T tarde · N noche · L libre · V vacaciones · B baja'
leg.font = Font(size=9, color='4A4A47'); leg.alignment = Alignment(vertical='center')

# Fila 5: número de día · Fila 6: letra del día de la semana (fórmula)
hdr_font = Font(bold=True, size=9, color=PAPER)
dow_font = Font(size=8, color='8A8A85')
ws.cell(row=5, column=1, value='Equipo').font = Font(bold=True, size=10, color=PAPER)
ws.cell(row=5, column=1).fill = HDR_FILL
ws.cell(row=5, column=2, value='').fill = HDR_FILL
ws.cell(row=6, column=1, value='')
ws.cell(row=6, column=2, value='')
for i in range(DAYS):
    col = FIRST_DAY_COL + i
    cl = get_column_letter(col)
    d = ws.cell(row=5, column=col)
    # días fuera de mes quedan vacíos
    d.value = f'=IF({i+1}<=DAY(EOMONTH(DATE($B$2,$B$3,1),0)),{i+1},"")'
    d.font = hdr_font; d.fill = HDR_FILL; d.alignment = Alignment(horizontal='center')
    w = ws.cell(row=6, column=col)
    w.value = f'=IF({cl}5="","",MID("LMXJVSD",WEEKDAY(DATE($B$2,$B$3,{cl}5),2),1))'
    w.font = dow_font; w.alignment = Alignment(horizontal='center')

# Columnas de recuento a la derecha
STAT_START = LAST_DAY_COL + 2  # una columna de aire
stats = [('Noches', 'N', TEAL_D), ('Findes', None, AMBER), ('Libres', 'L', '8A8A85')]
for j, (label, _, color) in enumerate(stats):
    c = ws.cell(row=5, column=STAT_START + j, value=label)
    c.font = Font(bold=True, size=8, color=color)
    c.alignment = Alignment(horizontal='center')
    ws.column_dimensions[get_column_letter(STAT_START + j)].width = 7

# Filas de personas
rng = lambda row: f'{get_column_letter(FIRST_DAY_COL)}{row}:{get_column_letter(LAST_DAY_COL)}{row}'
for k in range(NAMES):
    row = FIRST_PERSON_ROW + k
    name = ws.cell(row=row, column=1, value=f'Persona {k+1}')
    name.font = Font(size=10)
    name.border = BOX
    ws.cell(row=row, column=2, value='').border = BOX
    for i in range(DAYS):
        cell = ws.cell(row=row, column=FIRST_DAY_COL + i, value='')
        cell.alignment = Alignment(horizontal='center')
        cell.border = BOX
        cell.font = Font(size=10)
    # Recuentos
    n = ws.cell(row=row, column=STAT_START, value=f'=COUNTIF({rng(row)},"N")')
    # Findes: SUMPRODUCT sobre día de semana >5 y celda con turno de trabajo
    fst, lst = get_column_letter(FIRST_DAY_COL), get_column_letter(LAST_DAY_COL)
    wk = ws.cell(row=row, column=STAT_START + 1,
        value=f'=SUMPRODUCT(({fst}$5:{lst}$5<>"")*(WEEKDAY(DATE($B$2,$B$3,IF({fst}$5:{lst}$5="",1,{fst}$5:{lst}$5)),2)>5)*ISNUMBER(MATCH({fst}{row}:{lst}{row},{{"M";"T";"N"}},0)))')
    lb = ws.cell(row=row, column=STAT_START + 2, value=f'=COUNTIF({rng(row)},"L")+COUNTIF({rng(row)},"V")')
    for cst in (n, wk, lb):
        cst.alignment = Alignment(horizontal='center'); cst.font = Font(size=10)

# Bloque de cobertura
cov_label_row = LAST_PERSON_ROW + 2
ws.cell(row=cov_label_row, column=1, value='Cobertura por día').font = Font(bold=True, size=10, color=TEAL_D)
codes = ['M', 'T', 'N']
for j, code in enumerate(codes):
    row = cov_label_row + 1 + j
    ws.cell(row=row, column=1, value=f'Personas de {code}').font = Font(size=9, color='4A4A47')
    ws.cell(row=row, column=2, value='').border = BOX
    for i in range(DAYS):
        col = FIRST_DAY_COL + i
        cl = get_column_letter(col)
        c = ws.cell(row=row, column=col)
        c.value = f'=IF({cl}$5="","",COUNTIF({cl}${FIRST_PERSON_ROW}:{cl}${LAST_PERSON_ROW},"{code}"))'
        c.alignment = Alignment(horizontal='center'); c.font = Font(size=9); c.border = BOX

min_row = cov_label_row + 4 + 1
ws.cell(row=min_row, column=1, value='Mínimo por turno (M/T/N) →').font = Font(bold=True, size=9, color=INK)
for j, code in enumerate(codes):
    c = ws.cell(row=min_row, column=FIRST_DAY_COL + j, value=2 if code == 'N' else 3)
    c.fill = SOFT_FILL; c.border = BOX; c.alignment = Alignment(horizontal='center'); c.font = Font(size=9, bold=True)
    ws.cell(row=min_row, column=FIRST_DAY_COL + 3).value = '← edita estos tres valores'
    ws.cell(row=min_row, column=FIRST_DAY_COL + 3).font = Font(size=8, color='8A8A85')

# Formato condicional
person_range = f'{get_column_letter(FIRST_DAY_COL)}{FIRST_PERSON_ROW}:{get_column_letter(LAST_DAY_COL)}{LAST_PERSON_ROW}'
ws.conditional_formatting.add(person_range, CellIsRule(operator='equal', formula=['"N"'], fill=NIGHT_FILL))
for rest in ('"L"', '"V"', '"B"'):
    ws.conditional_formatting.add(person_range, CellIsRule(operator='equal', formula=[rest], fill=REST_FILL))
# Fin de semana sombreado en cabecera y cuadrícula
fst = get_column_letter(FIRST_DAY_COL)
wknd = f'AND({fst}$5<>"",WEEKDAY(DATE($B$2,$B$3,{fst}$5),2)>5)'
ws.conditional_formatting.add(f'{person_range}',
    FormulaRule(formula=[f'AND({fst}$5<>"",WEEKDAY(DATE($B$2,$B$3,{fst}$5),2)>5,{fst}{FIRST_PERSON_ROW}="")'], fill=PatternFill('solid', fgColor='F3F1EA')))
# Cobertura bajo mínimo en rojo
for j, code in enumerate(codes):
    row = cov_label_row + 1 + j
    cov_range = f'{get_column_letter(FIRST_DAY_COL)}{row}:{get_column_letter(LAST_DAY_COL)}{row}'
    min_cell = f'${get_column_letter(FIRST_DAY_COL + j)}${min_row}'
    ws.conditional_formatting.add(cov_range, FormulaRule(
        formula=[f'AND({fst}$5<>"",{fst}{row}<{min_cell})'],
        fill=PatternFill('solid', fgColor='F9E3E4'),
        font=Font(color=RED, bold=True)))

# Pie
foot_row = min_row + 3
ws.merge_cells(start_row=foot_row, start_column=1, end_row=foot_row, end_column=LAST_DAY_COL)
f = ws.cell(row=foot_row, column=1)
f.value = 'Esta plantilla cuenta y avisa, pero no puede vigilar descansos legales ni equidad. Auditoría gratuita del cuadrante: www.shiftia.es/recursos/auditoria-cuadrante'
f.font = Font(size=9, italic=True, color='8A8A85')

import os
os.makedirs('assets/downloads', exist_ok=True)
out = 'assets/downloads/plantilla-cuadrante-turnos-shiftia.xlsx'
wb.save(out)
print('OK', out)
